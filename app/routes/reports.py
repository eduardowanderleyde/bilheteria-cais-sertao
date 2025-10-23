"""Reports routes"""
from fastapi import APIRouter, Request, Depends, Query, HTTPException, status
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, desc, case
from datetime import datetime, date, timedelta
from typing import Optional
import pandas as pd
import tempfile
import os
from ..db import get_db
from ..models import Order, OrderItem, Group, GroupVisit
from ..auth import require_auth, get_user_info, can_export
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

# Helpers para mapeamento de dados
CASH_NAMES = {"cash", "dinheiro", "especie", "espécie"}
PIX_NAMES = {"pix"}
CARD_NAMES = {"cc", "cartao", "cartão", "credito", "crédito", "debit", "debito",
              "credit", "card", "visa", "master", "elo", "amex", "hipercard"}

def _pm_bucket(pm: str) -> str:
    s = (pm or "").strip().lower()
    if s in PIX_NAMES:
        return "pix"
    if s in CASH_NAMES:
        return "cash"
    # padrão: trata tudo como cartão (coluna CC)
    return "cc"

def _grat_bucket(reason: str):
    r = (reason or "").strip().upper()
    if r in {"DG", "DIA GRATUIDADE"}: return "DG"
    if r in {"GPD", "PCD"}: return "GPD"
    return "TG"

def _money(cents): return round((cents or 0)/100, 2)

def _to_date(d):
    if isinstance(d, date):        # inclui datetime.date
        return d
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(d, fmt).date()
            except ValueError:
                pass
    return None

thin = Side(style="thin")

def _write_bordero(ws, start_dt, end_dt, valores_ingresso, linhas):
    """
    valores_ingresso = {'inteira': 10.00, 'meia': 5.00}
    linhas = lista de dicts por dia já agregados
    """
    ws.title = "borderô"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("C6:U6")
    ws["C6"] = f"BILHETERIA CAIS DO SERTÃO - DETALHAMENTO DA MOVIMENTAÇÃO Nº / {end_dt.year} - GCS"
    ws["C6"].font = Font(bold=True, size=14)
    ws["C6"].alignment = Alignment(horizontal="center")

    # Box "Valor do ingresso"
    ws["C8"] = "VALOR DO INGRESSO"
    ws["C8"].font = Font(bold=True)
    ws["C9"], ws["D9"], ws["E9"] = "INTEIRA", f"R$ {valores_ingresso['inteira']:.2f}", ""
    ws["C10"], ws["D10"], ws["E10"] = "MEIA", f"R$ {valores_ingresso['meia']:.2f}", ""
    ws.merge_cells("C8:E8")

    # Cabeçalho principal da tabela (linha 12)
    base_row = 12
    headers = [
        ("DATA", 1),
        ("TIPOS DOS INGRESSOS / PAGAMENTO", 6),
        ("GRATUIDADE", 3),
        ("ARRECADAÇÃO", 4),
        ("PÚBLICO TOTAL DO DIA", 2),
    ]
    col = 3  # começa em C
    for text, span in headers:
        ws.merge_cells(start_row=base_row, start_column=col,
                       end_row=base_row, end_column=col+span-1)
        ws.cell(row=base_row, column=col, value=text).alignment = Alignment(horizontal="center")
        ws.cell(row=base_row, column=col).font = Font(bold=True)
        col += span

    # Subcabeçalhos (linha 13)
    sub_labels = [
        "DATA",
        "$$","PIX","CC","$$","PIX","CC",  # INTEIRA e MEIA
        "DG","GPD","TG",
        "VALOR EM ESPECIE","VALOR PIX","VALOR CC","RECEITA DO DIA(R$)",
        "PAGANTES DO DIA","P+G+E DO DIA"
    ]
    
    # Escrever subcabeçalhos
    for i, label in enumerate(sub_labels, start=3):
        ws.cell(row=base_row+1, column=i, value=label)
        ws.cell(row=base_row+1, column=i).font = Font(bold=True)
        ws.cell(row=base_row+1, column=i).alignment = Alignment(horizontal="center")

    # Larguras das colunas
    widths = [12] + [8]*6 + [8,8,8] + [16,12,10,18] + [16,16]
    for i, w in enumerate(widths, start=3):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

    # Período
    ws["C11"] = "ADM. EMPETUR / PERÍODO"
    ws["C11"].font = Font(bold=True)
    ws["D11"] = f"DE {start_dt:%d.%m} A {end_dt:%d.%m.%Y}"

    # Linhas de dados (começa na linha 14)
    r = base_row + 2
    for linha in linhas:
        day = _to_date(linha["date"])
        ws.cell(row=r, column=3, value=day.strftime("%d.%m") if day else str(linha["date"] or ""))
        ws.cell(row=r, column=4, value=int(linha["qtd_int_cash"]))
        ws.cell(row=r, column=5, value=int(linha["qtd_int_pix"]))
        ws.cell(row=r, column=6, value=int(linha["qtd_int_cc"]))
        ws.cell(row=r, column=7, value=int(linha["qtd_meia_cash"]))
        ws.cell(row=r, column=8, value=int(linha["qtd_meia_pix"]))
        ws.cell(row=r, column=9, value=int(linha["qtd_meia_cc"]))
        ws.cell(row=r, column=10, value=int(linha["g_DG"]))
        ws.cell(row=r, column=11, value=int(linha["g_GPD"]))
        ws.cell(row=r, column=12, value=int(linha["g_TG"]))
        ws.cell(row=r, column=13, value=_money(linha["rec_cash"]))
        ws.cell(row=r, column=14, value=_money(linha["rec_pix"]))
        ws.cell(row=r, column=15, value=_money(linha["rec_cc"]))
        ws.cell(row=r, column=16, value=_money(linha["rec_cash"] + linha["rec_pix"] + linha["rec_cc"]))
        ws.cell(row=r, column=17, value=int(linha["pagantes"]))
        ws.cell(row=r, column=18, value=int(linha["publico_total"]))
        r += 1

    # Totais do período
    ws.cell(row=r, column=3, value="TOTAIS DA SEMANA").font = Font(bold=True)
    for col_idx in range(4, 19):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.cell(row=r, column=col_idx, value=f"=SUM({col_letter}{base_row+2}:{col_letter}{r-1})").font = Font(bold=True)

    # Assinaturas
    r += 3
    ws.cell(row=r, column=3, value="EXECUTIVO SÊNIOR")
    ws.cell(row=r, column=9, value="VISTO GESTOR")
    r += 2
    ws.cell(row=r, column=3, value="_________________")
    ws.cell(row=r, column=9, value="_________________")
    r += 1
    ws.cell(row=r, column=3, value="")
    ws.cell(row=r, column=9, value="")

    # Bordas nas linhas de dados
    for rr in range(base_row, r):
        for cc in range(3, 19):
            ws.cell(row=rr, column=cc).border = Border(top=thin, left=thin, right=thin, bottom=thin)

router = APIRouter()
templates = Jinja2Templates(directory="templates")

@router.get("/", response_class=HTMLResponse)
async def reports_page(request: Request):
    """Reports page"""
    user = require_auth(request)
    
    return templates.TemplateResponse("reports.html", {
        "request": request,
        "user": get_user_info(request)
    })

@router.get("/reasons", response_class=HTMLResponse)
async def reports_reasons_page(request: Request):
    """Reports by discount reasons page"""
    user = require_auth(request)
    
    return templates.TemplateResponse("reports_reasons.html", {
        "request": request,
        "user": get_user_info(request)
    })

@router.get("/groups", response_class=HTMLResponse)
async def reports_groups_page(request: Request):
    """Reports groups page"""
    user = require_auth(request)
    
    return templates.TemplateResponse("reports_groups.html", {
        "request": request,
        "user": get_user_info(request)
    })

@router.get("/api/export")
async def reports_export(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """General reports export (Excel)"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Query daily data
    daily_results = db.query(
        func.date(Order.created_at).label('date'),
        func.sum(OrderItem.qty).label('total_people'),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label('total_revenue')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at)).order_by('date').all()
    
    # Query by ticket type
    type_results = db.query(
        func.date(Order.created_at).label('date'),
        OrderItem.ticket_type,
        func.sum(OrderItem.qty).label('count')
    ).join(Order).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at), OrderItem.ticket_type).all()
    
    # Query by payment method
    payment_results = db.query(
        func.date(Order.created_at).label('date'),
        Order.payment_method,
        func.sum(OrderItem.qty).label('count')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at), Order.payment_method).all()
    
    # Create DataFrames
    daily_df = pd.DataFrame([
        {
            'Data': result.date,
            'Pessoas': result.total_people,
            'Receita (R$)': round(result.total_revenue / 100, 2)
        }
        for result in daily_results
    ])
    
    type_df = pd.DataFrame([
        {
            'Data': result.date,
            'Tipo': result.ticket_type,
            'Quantidade': result.count
        }
        for result in type_results
    ])
    
    payment_df = pd.DataFrame([
        {
            'Data': result.date,
            'Forma de Pagamento': result.payment_method,
            'Quantidade': result.count
        }
        for result in payment_results
    ])
    
    # Create Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            daily_df.to_excel(writer, sheet_name='Resumo_Diario', index=False)
            type_df.to_excel(writer, sheet_name='Por_Tipo', index=False)
            payment_df.to_excel(writer, sheet_name='Por_Pagamento', index=False)
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"relatorio_geral_{start_dt}_{end_dt}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.get("/by-state")
async def report_by_state(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Report by state (CSV)"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Query data
    results = db.query(
        Order.state,
        func.sum(OrderItem.qty).label('total_people'),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label('total_revenue')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(Order.state).order_by(desc('total_people')).all()
    
    # Create DataFrame
    df = pd.DataFrame([
        {
            'UF': result.state or 'Não informado',
            'Pessoas': result.total_people,
            'Receita (R$)': round(result.total_revenue / 100, 2)
        }
        for result in results
    ])
    
    # Create temporary file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        df.to_csv(tmp.name, index=False, encoding='utf-8')
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"pessoas_por_uf_{start_dt}_{end_dt}.csv",
        media_type="text/csv"
    )

@router.get("/by-discount-reason")
async def report_by_discount_reason(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Report by discount reason (CSV)"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Query data
    results = db.query(
        OrderItem.discount_reason,
        func.sum(OrderItem.qty).label('count'),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label('total_revenue')
    ).join(Order).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None),
            OrderItem.discount_reason.isnot(None)
        )
    ).group_by(OrderItem.discount_reason).order_by(desc('count')).all()
    
    # Create DataFrame
    df = pd.DataFrame([
        {
            'Motivo': result.discount_reason or 'Não informado',
            'Quantidade': result.count,
            'Receita (R$)': round(result.total_revenue / 100, 2)
        }
        for result in results
    ])
    
    # Create temporary file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        df.to_csv(tmp.name, index=False, encoding='utf-8')
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"motivos_desconto_{start_dt}_{end_dt}.csv",
        media_type="text/csv"
    )

@router.get("/by-payment-method")
async def report_by_payment_method(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Report by payment method (CSV)"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Query data
    results = db.query(
        Order.payment_method,
        func.sum(OrderItem.qty).label('count'),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label('total_revenue')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(Order.payment_method).order_by(desc('total_revenue')).all()
    
    # Create DataFrame
    df = pd.DataFrame([
        {
            'Forma de Pagamento': result.payment_method,
            'Quantidade': result.count,
            'Receita (R$)': round(result.total_revenue / 100, 2)
        }
        for result in results
    ])
    
    # Create temporary file
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
        df.to_csv(tmp.name, index=False, encoding='utf-8')
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"formas_pagamento_{start_dt}_{end_dt}.csv",
        media_type="text/csv"
    )

@router.get("/by-payment.csv")
async def report_by_payment_csv(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Export payment method report as CSV"""
    user = require_auth(request)
    
    # Parse dates
    start = None
    end = None
    if start_date:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
    if end_date:
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
    
    # Build query
    query = db.query(
        Order.payment_method,
        func.count(Order.id).label("qtd"),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label("total_cents")
    ).join(OrderItem).filter(Order.deleted_at.is_(None))
    
    if start:
        query = query.filter(func.date(Order.created_at) >= start)
    if end:
        query = query.filter(func.date(Order.created_at) <= end)
    
    results = query.group_by(Order.payment_method).all()
    
    # Create DataFrame
    data = []
    for result in results:
        data.append({
            "forma_pagamento": result.payment_method,
            "quantidade_vendas": result.qtd,
            "receita_total": round((result.total_cents or 0) / 100, 2)
        })
    
    df = pd.DataFrame(data)
    
    # Create temporary file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode='w', encoding='utf-8')
    df.to_csv(tmp.name, index=False)
    tmp.close()
    
    return FileResponse(
        tmp.name,
        filename="vendas_por_pagamento.csv",
        media_type="text/csv"
    )

@router.get("/daily")
async def daily_report(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Daily report (Excel)"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Query daily data
    daily_results = db.query(
        func.date(Order.created_at).label('date'),
        func.sum(OrderItem.qty).label('total_people'),
        func.sum(OrderItem.qty * OrderItem.unit_price_cents).label('total_revenue')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at)).order_by('date').all()
    
    # Query by ticket type
    type_results = db.query(
        func.date(Order.created_at).label('date'),
        OrderItem.ticket_type,
        func.sum(OrderItem.qty).label('count')
    ).join(Order).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at), OrderItem.ticket_type).all()
    
    # Query by payment method
    payment_results = db.query(
        func.date(Order.created_at).label('date'),
        Order.payment_method,
        func.sum(OrderItem.qty).label('count')
    ).join(OrderItem).filter(
        and_(
            func.date(Order.created_at) >= start_dt,
            func.date(Order.created_at) <= end_dt,
            Order.deleted_at.is_(None)
        )
    ).group_by(func.date(Order.created_at), Order.payment_method).all()
    
    # Create DataFrames
    daily_df = pd.DataFrame([
        {
            'Data': result.date,
            'Pessoas': result.total_people,
            'Receita (R$)': round(result.total_revenue / 100, 2)
        }
        for result in daily_results
    ])
    
    type_df = pd.DataFrame([
        {
            'Data': result.date,
            'Tipo': result.ticket_type,
            'Quantidade': result.count
        }
        for result in type_results
    ])
    
    payment_df = pd.DataFrame([
        {
            'Data': result.date,
            'Forma de Pagamento': result.payment_method,
            'Quantidade': result.count
        }
        for result in payment_results
    ])
    
    # Create Excel file
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            daily_df.to_excel(writer, sheet_name='Resumo_Diario', index=False)
            type_df.to_excel(writer, sheet_name='Por_Tipo', index=False)
            payment_df.to_excel(writer, sheet_name='Por_Pagamento', index=False)
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"relatorio_diario_{start_dt}_{end_dt}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Group reports endpoints
def _period_days(days: int):
    start = date.today() - timedelta(days=days)
    return start, date.today()

@router.get("/api/groups/weekly")
def groups_weekly(db: Session = Depends(get_db), weeks: int = 8):
    """Weekly groups report"""
    # Cap maximum weeks
    weeks = min(weeks, 52)  # Max 1 year
    
    start = date.today() - timedelta(days=7*weeks)
    
    # Use SQLite strftime for compatibility
    q = (db.query(
            func.strftime('%Y-%W', GroupVisit.date).label('year_week'),
            func.count(GroupVisit.id).label('groups'),
            func.coalesce(func.sum(GroupVisit.size), 0).label('people')
         )
         .filter(GroupVisit.date >= start)
         .group_by(func.strftime('%Y-%W', GroupVisit.date))
         .order_by(func.strftime('%Y-%W', GroupVisit.date)))
    
    return [{"bucket": r[0], "groups": int(r[1]), "people": int(r[2])} for r in q.all()]

@router.get("/api/groups/monthly")
def groups_monthly(db: Session = Depends(get_db), months: int = 12):
    """Monthly groups report"""
    # Cap maximum months
    months = min(months, 24)  # Max 2 years
    
    start = date.today().replace(day=1) - timedelta(days=30*months)
    
    # Use SQLite strftime for compatibility (PostgreSQL would use date_trunc)
    q = (db.query(
            func.strftime('%Y-%m', GroupVisit.date).label('ym'),
            func.count(GroupVisit.id).label('groups'),
            func.coalesce(func.sum(GroupVisit.size), 0).label('people')
         )
         .filter(GroupVisit.date >= start)
         .group_by(func.strftime('%Y-%m', GroupVisit.date))
         .order_by(func.strftime('%Y-%m', GroupVisit.date)))
    
    return [{"month": r[0], "groups": int(r[1]), "people": int(r[2])} for r in q.all()]

@router.get("/api/groups/top-origins")
def groups_top_origins(db: Session = Depends(get_db), days: int = 180, limit: int = 10):
    """Top origins for groups"""
    # Cap maximum values
    days = min(days, 365)  # Max 1 year
    limit = min(limit, 100)  # Max 100 results
    
    start, _ = _period_days(days)
    q = (db.query(GroupVisit.state, GroupVisit.city,
                  func.count(GroupVisit.id).label('groups'),
                  func.coalesce(func.sum(GroupVisit.size), 0).label('people'))
           .filter(GroupVisit.date >= start)
           .group_by(GroupVisit.state, GroupVisit.city)
           .order_by(desc('people')).limit(limit))
    return [{"state": s or "", "city": c or "", "groups": int(g), "people": int(p)} for s,c,g,p in q.all()]

@router.get("/api/groups/kpis")
def groups_kpis(db: Session = Depends(get_db), days: int = 30):
    """Groups KPIs"""
    # Cap maximum days
    days = min(days, 365)  # Max 1 year
    
    start, _ = _period_days(days)
    
    # Single query for all KPIs (more efficient)
    result = db.query(
        func.count(GroupVisit.id).label('groups'),
        func.coalesce(func.sum(GroupVisit.size), 0).label('people'),
        func.count(func.distinct(GroupVisit.date)).label('days_with')
    ).filter(GroupVisit.date >= start).first()
    
    return {
        "groups": int(result.groups or 0), 
        "people": int(result.people or 0), 
        "days_with": int(result.days_with or 0)
    }

@router.get("/groups/export.xlsx")
def groups_export(db: Session = Depends(get_db), months: int = 12):
    """Export groups to Excel with multiple sheets"""
    # Cap maximum months
    months = min(months, 24)  # Max 2 years
    
    start_date = date.today().replace(day=1) - timedelta(days=30*months)
    
    # 1) Raw data (limit to prevent memory issues)
    rows = (db.query(GroupVisit.date, GroupVisit.institution, GroupVisit.size,
                     GroupVisit.state, GroupVisit.city, GroupVisit.scheduled, GroupVisit.price_total)
              .filter(GroupVisit.date >= start_date)
              .order_by(GroupVisit.date.desc())
              .limit(10000)  # Cap at 10k records
              .all())
    df_raw = pd.DataFrame(rows, columns=["Data","Instituição","Pessoas","UF","Cidade","Agendada","ValorTotal"])

    # 2) Monthly aggregation (SQL-based for performance)
    monthly_rows = (db.query(
        func.strftime('%Y-%m', GroupVisit.date).label('month'),
        func.count(GroupVisit.id).label('grupos'),
        func.coalesce(func.sum(GroupVisit.size), 0).label('pessoas'),
        func.coalesce(func.sum(GroupVisit.price_total), 0).label('valor_total')
    )
    .filter(GroupVisit.date >= start_date)
    .group_by(func.strftime('%Y-%m', GroupVisit.date))
    .order_by(func.strftime('%Y-%m', GroupVisit.date))
    .all())
    
    df_month = pd.DataFrame(monthly_rows, columns=["Mês", "Grupos", "Pessoas", "ValorTotal"])

    # 3) Weekly aggregation (SQL-based)
    weekly_rows = (db.query(
        func.strftime('%Y-%W', GroupVisit.date).label('week'),
        func.count(GroupVisit.id).label('grupos'),
        func.coalesce(func.sum(GroupVisit.size), 0).label('pessoas'),
        func.coalesce(func.sum(GroupVisit.price_total), 0).label('valor_total')
    )
    .filter(GroupVisit.date >= start_date)
    .group_by(func.strftime('%Y-%W', GroupVisit.date))
    .order_by(func.strftime('%Y-%W', GroupVisit.date))
    .all())
    
    df_week = pd.DataFrame(weekly_rows, columns=["Semana", "Grupos", "Pessoas", "ValorTotal"])

    # 4) Top origins (SQL-based)
    origin_rows = (db.query(
        GroupVisit.state.label('uf'),
        GroupVisit.city.label('cidade'),
        func.count(GroupVisit.id).label('grupos'),
        func.coalesce(func.sum(GroupVisit.size), 0).label('pessoas')
    )
    .filter(GroupVisit.date >= start_date)
    .group_by(GroupVisit.state, GroupVisit.city)
    .order_by(desc('pessoas'))
    .limit(50)
    .all())
    
    df_origin = pd.DataFrame(origin_rows, columns=["UF", "Cidade", "Grupos", "Pessoas"])

    # Create Excel file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as xw:
        df_raw.to_excel(xw, index=False, sheet_name="Bruto")
        df_month.to_excel(xw, index=False, sheet_name="Mensal")
        df_week.to_excel(xw, index=False, sheet_name="Semanal")
        df_origin.to_excel(xw, index=False, sheet_name="TopOrigens")

    return FileResponse(
        tmp.name, 
        filename="Relatorio_Grupos.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.get("/groups/export.csv")
def groups_export_csv(db: Session = Depends(get_db), months: int = 12):
    """Export groups to CSV (streaming for large datasets)"""
    # Cap maximum months
    months = min(months, 24)  # Max 2 years
    
    start_date = date.today().replace(day=1) - timedelta(days=30*months)
    
    def generate_csv():
        """Generator for CSV data"""
        # Header
        yield "Data,Instituição,Pessoas,UF,Cidade,Agendada,ValorTotal\n"
        
        # Data rows (streaming)
        query = (db.query(GroupVisit.date, GroupVisit.institution, GroupVisit.size,
                         GroupVisit.state, GroupVisit.city, GroupVisit.scheduled, GroupVisit.price_total)
                  .filter(GroupVisit.date >= start_date)
                  .order_by(GroupVisit.date.desc())
                  .yield_per(1000))  # Process in chunks
        
        for row in query:
            # Escape CSV values
            date_str = row.date.strftime("%Y-%m-%d") if row.date else ""
            institution = str(row.institution or "").replace(",", ";").replace("\n", " ")
            size = str(row.size or 0)
            state = str(row.state or "")
            city = str(row.city or "").replace(",", ";").replace("\n", " ")
            scheduled = "Sim" if row.scheduled else "Não"
            price = str(row.price_total or 0)
            
            yield f"{date_str},{institution},{size},{state},{city},{scheduled},{price}\n"
    
    return StreamingResponse(
        generate_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=grupos.csv"}
    )

@router.get("/bordero-cais")
async def bordero_cais(
    request: Request,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Borderô Cais - Relatório consolidado completo"""
    user = require_auth(request)
    
    if not can_export(user["role"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions"
        )
    
    # Parse dates - default to last 30 days
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today() - timedelta(days=30)
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else date.today()
    
    # Agregação por dia para o borderô
    rows = (db.query(
                func.date(Order.created_at).label("d"),
                OrderItem.ticket_type,
                Order.payment_method,
                func.coalesce(func.sum(OrderItem.qty), 0).label("qtd"),
                func.coalesce(func.sum(OrderItem.qty * OrderItem.unit_price_cents), 0).label("cents"),
                func.max(OrderItem.discount_reason).label("discount_reason"),
                func.max(case((OrderItem.unit_price_cents == 0, 1), else_=0)).label("has_free")
           )
           .select_from(Order)
           .join(OrderItem, OrderItem.order_id == Order.id)
           .filter(Order.deleted_at.is_(None),
                   Order.created_at >= start_dt,
                   Order.created_at < (end_dt + timedelta(days=1)))
           .group_by(func.date(Order.created_at), OrderItem.ticket_type, Order.payment_method)
           .all())

    by_day = defaultdict(lambda: {
        "q_int": {"cash":0,"pix":0,"cc":0},
        "q_meia":{"cash":0,"pix":0,"cc":0},
        "g":{"DG":0,"GPD":0,"TG":0},
        "rec":{"cash":0,"pix":0,"cc":0},
    })
    
    # Preencher dados por dia
    for r in rows:
        day = _to_date(r.d)
        pm = _pm_bucket(r.payment_method)
        tt = (r.ticket_type or "").lower()
        if (r.cents or 0) == 0:
            by_day[day]["g"][_grat_bucket(r.discount_reason)] += int(r.qtd or 0)
        else:
            if tt == "inteira":
                by_day[day]["q_int"].setdefault(pm, 0)
                by_day[day]["q_int"][pm] += int(r.qtd or 0)
            elif tt == "meia":
                by_day[day]["q_meia"].setdefault(pm, 0)
                by_day[day]["q_meia"][pm] += int(r.qtd or 0)
            by_day[day]["rec"].setdefault(pm, 0)
            by_day[day]["rec"][pm] += int(r.cents or 0)

    # Montar linhas para o borderô
    linhas = []
    for d in sorted(by_day.keys()):
        b = by_day[d]
        pagantes = sum(b["q_int"].values()) + sum(b["q_meia"].values())
        publico = pagantes + b["g"]["DG"] + b["g"]["GPD"] + b["g"]["TG"]
        linhas.append({
            "date": d,
            "qtd_int_cash": b["q_int"]["cash"],
            "qtd_int_pix":  b["q_int"]["pix"],
            "qtd_int_cc":   b["q_int"]["cc"],
            "qtd_meia_cash": b["q_meia"]["cash"],
            "qtd_meia_pix":  b["q_meia"]["pix"],
            "qtd_meia_cc":   b["q_meia"]["cc"],
            "g_DG":  b["g"]["DG"],
            "g_GPD": b["g"]["GPD"],
            "g_TG":  b["g"]["TG"],
            "rec_cash": b["rec"]["cash"],
            "rec_pix":  b["rec"]["pix"],
            "rec_cc":   b["rec"]["cc"],
            "pagantes": pagantes,
            "publico_total": publico,
        })
    
    # Criar Borderô Cais - Relatório Consolidado
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            # Criar aba do borderô
            ws = writer.book.create_sheet("borderô")
            _write_bordero(ws, start_dt, end_dt, {"inteira": 10.00, "meia": 5.00}, linhas)
        
        tmp_path = tmp.name
    
    return FileResponse(
        tmp_path,
        filename=f"Borderô_Cais_{start_dt}_{end_dt}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )